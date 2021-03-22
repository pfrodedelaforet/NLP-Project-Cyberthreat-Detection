import openpyxl
import unshortenit
from unshortenit import UnshortenIt
unshortener = UnshortenIt()
wb = openpyxl.load_workbook("tablerase.xlsx")
ws = wb.active
n_tweets = 5332
for i in range(1, n_tweets+1):
    try :
        link = ws["D{}".format(i)].value
        ws["D{}".format(i)] = unshortener.unshorten(link)
    except :
        ws["D{}".format(i)] = "lien non valide..."
wb.save('tablerase_2.xlsx')