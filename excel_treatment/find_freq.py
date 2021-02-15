import openpyxl
from datetime import datetime, timedelta
import re
wb = openpyxl.load_workbook("new_dataset_inv_search.xlsx")
ws = wb["Sheet1"]
ws3 = wb.create_sheet("days_per_tweet")
username = ws["B1"].value
dico = dict()
j=1
for i in range(1, 4004):
    if ws[f"B{i}"].value != username and j!=i-1:
        d1=ws[f"C{j}"].value
        d2 = ws[f"C{i-1}"].value
        date1 = datetime(*[int(x) for x in re.split('-|:| ', d1)])
        date2 = datetime(*[int(x) for x in re.split('-|:| ', d2)])
        print(j, i-1)
        dico[username] = (date1-date2).days/(i-1-j)
        print(dico[username])
        j=i
        username = ws[f"B{i}"].value
i = 2
for user in dico :
    ws3[f"A{i}"] = user
    ws3[f"B{i}"] = dico[user]
    i += 1
wb.save('new_dataset_inv_search_freq.xlsx')