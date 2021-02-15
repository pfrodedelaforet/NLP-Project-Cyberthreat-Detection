import tweepy as tw
import xlsxwriter

import openpyxl
#permet de récupérer les mentions à partir des articles initiaux
wb = openpyxl.load_workbook("nouveaux_ids_derniere_v_2.xlsx")
ws = wb["Sheet1"]
with open('cles.txt', 'r') as f:
    lignes = []
    for line in f:
        l = line.strip().split(' : ')
        lignes.append(l)
credentials = {}
credentials[lignes[0][0]] = lignes[0][1]
credentials[lignes[1][0]] = lignes[1][1]
credentials[lignes[2][0]] = lignes[2][1]
credentials[lignes[3][0]] = lignes[3][1]

auth = tw.OAuthHandler(credentials['CONSUMER_KEY'], credentials['CONSUMER_SECRET'])
auth.set_access_token(credentials['ACCESS_TOKEN'], credentials['ACCESS_SECRET'])
ws = wb.active
ws2 = wb.create_sheet('inv_search_mentions')
API = tw.API(auth, wait_on_rate_limit=True)
k = 1
for i in range(1, 550):
    if ws[f"H{i}"].value == 'RELEVANT':
        id = ws[f"A{i}"].value
        link = ws[f"D{i}"].value
        try :
            stat = API.get_status(id)
            screen_name = stat.user.screen_name
            link_tweet = 'https://twitter.com/'+screen_name+'/status/'+id
            print(link_tweet)
            list_tweets = API.search(link_tweet)
            if len(list_tweets) != 0 :
                print('oui')
                for tweet in list_tweets :
                    ws2["A{}".format(k)] = str(tweet.id)
                    ws2["B{}".format(k)] = tweet.user.screen_name
                    ws2["C{}".format(k)] = str(tweet.created_at)
                    ws2["D{}".format(k)] = link
                    ws2["E{}".format(k)] = tweet.retweet_count
                    ws2["F{}".format(k)] = tweet.favorite_count
                    ws2["G{}".format(k)] = tweet.text
                    k += 1
            else :
                print('non')
        except :
            print(f'pas de statut trouvé à i={i}')

wb.save('essai_mentions_2.xlsx')