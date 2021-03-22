import openpyxl
import csv
from tqdm import tqdm
"""This script allows to convert the excel sheet containing the dataset to folders relevant/irrelevant containing text files."""
wb = openpyxl.load_workbook("../storage/dataset/new_dataset_inv_search_freq.xlsx")
ws = wb["Sheet1"]
n_tweets = 3556
for i in tqdm(range(1, n_tweets+1)):
    num = str(i)
    tweet_id = str(ws[f"A{i}"].value)
    print(tweet_id)
    username = str(ws[f"B{i}"].value)
    article_text = str(ws[f"J{i}"].value)
    labels = str(ws[f"H{i}"].value)
    file = open(f'../storage/dataset/articles/{labels}/article_{num}_{tweet_id}_{username}.txt', 'w', encoding='utf-8')
    print(article_text, file = file)
file.close()
