
import tweepy as tw
import openpyxl
wb = openpyxl.load_workbook("copie.xlsx")
#ca devait être pour changer les anciens ids et remettre les contenus de tweet sous forme lisible
ws = wb["Sheet1"]
ws2 = wb['inv_search']
with open('cles.txt', 'r') as f:
    lignes = []
    for line in f:
        l = line.strip().split(' : ')
        lignes.append(l)
credentials = {}
credentials[lignes[0][0]] = lignes[0][1]
credentials[lignes[1][0]] = lignes[1][1]
credentials[lignes[2][0]] = lignes[2][1]
credentials[lignes[3][0]] = lignes[3][1]

auth = tw.OAuthHandler(credentials['CONSUMER_KEY'], credentials['CONSUMER_SECRET'])
auth.set_access_token(credentials['ACCESS_TOKEN'], credentials['ACCESS_SECRET'])
API = tw.API(auth, wait_on_rate_limit=True)
contenu = "RT @JAMESWT_MHT: #Rat malware from Steganography image"
liste_inconnus = []
for i in range(1, 528):
    p = True ; contenu = ws["G{}".format(i)].value ; screen_name = ws["B{}".format(i)].value ; old_id = str(ws["A{}".format(i)].value)
    contenu = contenu[:min(len(contenu), 120)]
    if contenu[0:2] == "RT":
        s = 0
        print('a')
        while contenu[s] != ":":
            s+= 1
        screen_name = contenu[4:s]
        contenu = contenu[s+2:]
    res = API.user_timeline(screen_name = screen_name, count = 500)
    p = True
    if 'http' in contenu :
    print(contenu)
    for r in res :
        if contenu in r.text:
            ws["A{}".format(i)] = str(r.id)
            p = False
            break
    if p :
        liste_inconnus.append(i)
    print(p)

wb.save('nouveaux_ids_derniere_v.xlsx')