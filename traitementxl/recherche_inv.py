import tweepy as tw
import xlsxwriter

import openpyxl
wb = openpyxl.load_workbook("new_dataset_while_labelling_8.xlsx")

with open('cles.txt', 'r') as f:
    lignes = []
    for line in f:
        l = line.strip().split(' : ')
        lignes.append(l)
credentials = {}
credentials[lignes[0][0]] = lignes[0][1]
credentials[lignes[1][0]] = lignes[1][1]
credentials[lignes[2][0]] = lignes[2][1]
credentials[lignes[3][0]] = lignes[3][1]

auth = tw.OAuthHandler(credentials['CONSUMER_KEY'], credentials['CONSUMER_SECRET'])
auth.set_access_token(credentials['ACCESS_TOKEN'], credentials['ACCESS_SECRET'])
ws = wb.active
ws2 = wb.create_sheet('inv_search')
API = tw.API(auth, wait_on_rate_limit=True)
links = set()#links is the set of relevant links counted once per link
k = 2#compteur des lignes remplies
for i in range(1, 4003):
    if ws["H{}".format(i)].value == 'RELEVANT':
        link = ws["D{}".format(i)].value
        links.add(link)
print(len(links))
for link in links:
    print(link)
    search_res = API.search(link)

    print(len(search_res))
    for i in range(len(search_res)):
        tweet = search_res[i]
        ws2["A{}".format(k)] = tweet.id
        ws2["B{}".format(k)] = tweet.user.screen_name
        ws2["C{}".format(k)] = tweet.created_at
        ws2["D{}".format(k)] = link
        ws2["E{}".format(k)] = tweet.retweet_count
        ws2["F{}".format(k)] = tweet.favorite_count
        ws2["G{}".format(k)] = tweet.text
        k += 1

wb.save('new_dataset_inv_search.xlsx')